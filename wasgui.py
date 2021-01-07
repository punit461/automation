from tkinter import *
from openpyxl import Workbook
from tkinter import messagebox
from tkinter.filedialog import askopenfile
import os
import webbrowser
from pathlib import Path
import shutil


def open_sheets():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Number"
    sheet["C1"] = "Message"
    sheet["E2"] = "Use ALT + Enter to go to New Line Inside the Cell and " \
                  "Concatinate function to add 2 or Multiple cells "
    sheet["E3"] = "And Dont Worry you can Delete this Instructions and use the Space" \
                  "but Dont change the Name, Num, message and input accordingly "
    workbook.save(filename="wasgui.xlsx")
    os.startfile("wasgui.xlsx")

    source = 'C:\\Codes_projects_python\\WASGUI\\Attachment\\'
    dest1 = 'C:\\Codes_projects_python\\WASGUI\\image\\'

    files = os.listdir(source)

    for f in files:
        shutil.move(source + f, dest1)


def attach():
    import os
    path = "C:\\link\\TEST1\\TEST\\Attachment"
    path = os.path.realpath(path)
    os.startfile(path)


class no_command():

    def __init__(self):
        self.btnclk = 0

    def no_option(self):
        no_command.__init__(self)
        messagebox.showinfo("Attachment", "You Have not selected any attachment only text message will be sent")
        self.btnclk += 1
        print(self.btnclk)


def wasend():
    # check the Total Count Thru z Already the formula in sheet
    z = int(sheet.cell(1, 6).value)
    print(z)

    # Assign Initial ref
    a = 2
    b = 1
    temp = 0
    # while loop to control the infinite loop in the program
    while a <= z:
        if temp <= 20:
            #nam = sheet.cell(a, 1).value

            num = sheet.cell(a, 2).value
            mes = sheet.cell(a, 3).value
            url = "https://web.whatsapp.com/send?phone=91" + num + "&text=" + mes + "&source&data=0&app_absent=1"
            driver.get(url)
            # driver.find_element_by_xpath("/html/body/div[1]/div/div/div[4]/div/footer/div[1]/div[2]/div/div[2]").clear()
            # driver.find_element_by_xpath("/html/body/div[1]/div/div/div[4]/div/footer/div[1]/div[2]/div/div[2]").send_keys(mes)
            driver.implicitly_wait(50)
            driver.find_element_by_class_name("_35EW6").click()
            temp += 1
            a += 1
        else:
            time.sleep(30)
            temp = 0


Nobtn = no_command()
root = Tk()
root.title('WASGUI')
topFrame = Frame(root)
topFrame.pack(side=TOP)

middleFrame = Frame(root)
middleFrame.pack()

bottomFrame = Frame(root)
bottomFrame.pack()

w1 = Label(topFrame, text='Enter Data In the Sheet!', height=2)
w1.pack()

button1 = Button(topFrame, text='Open Sheet', width=22, command=lambda: open_sheets())
button1.pack()

w2 = Label(middleFrame, text='Add Attachment?', height=2)
w2.pack()
button2a = Button(middleFrame, text='Yes', width=10, fg='green', command=lambda: attach())
button2a.pack(side='left')
button2b = Button(middleFrame, text='No', width=10, fg='red', command=lambda: Nobtn.no_option())
button2b.pack(side='right')

w3 = Label(bottomFrame, text='Press button to Start or Stop', height=2)
w3.pack()

button3a = Button(bottomFrame, text='Start', width=10, fg='green', command=lambda: wasend())
button3a.pack(side='left')
button3b = Button(bottomFrame, text='Stop', width=10, fg='red', command=root.destroy)
button3b.pack(side='right')

root.minsize(250, 200)
root.mainloop()
